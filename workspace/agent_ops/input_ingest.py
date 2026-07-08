# -*- coding: utf-8 -*-
"""Input ingestion: read the user's actual files so artifacts are grounded.

Real office/engineering requests come with material — test-result CSVs, mail
lists, existing macros, logs, draft documents. This module turns those into a
structured, secret-free summary (facts, notable/abnormal items, parsed mails)
that the planner and artifact generators consume, so a generated report cites
the actual file, row counts, and failures instead of generic TODOs.

Design rules:
  - stdlib-only; Korean/space paths supported (all paths go through pathlib)
  - explicit user-supplied paths only; directories are scanned with limits
  - large files: only the first MAX_BYTES_FULL bytes are analyzed (noted)
  - binary/unknown formats are recorded as unsupported, never silently skipped
  - secret-like lines (password/api key/token=...) are masked before any
    fact/preview leaves this module — summaries land in artifacts/diagnostics
"""
from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from .encoding_ops import decode_file_bytes
from . import doc_convert

try:
    import openpyxl  # type: ignore
except ImportError:  # optional dependency; xlsx degrades to unsupported
    openpyxl = None  # type: ignore

SUPPORTED_SUFFIXES = {".md", ".txt", ".csv", ".tsv", ".xlsx", ".log", ".py", ".bas", ".bat", ".json"}
MAX_BYTES_FULL = 200_000   # analyze at most this much per file
MAX_FILES = 50             # directory scan cap
_PREVIEW_CHARS = 300

_SECRET_LINE = re.compile(
    r"(password|passwd|secret|api[_-]?key|token|authorization)\s*[:=]", re.IGNORECASE)
# Row-level markers that make a data row worth surfacing to the user.
_ABNORMAL_MARKERS = ("불합격", "이상", "초과", "미달", "불량", "fail", "ng", "error")
_ENTRY_RE = re.compile(
    r"^\s*(?:Public\s+|Private\s+)?(?:Sub|Function|def)\s+([A-Za-z0-9_가-힣]+)",
    re.MULTILINE)


def _mask_secrets(text: str) -> str:
    return "\n".join(
        "[masked: secret-like line]" if _SECRET_LINE.search(line) else line
        for line in text.splitlines())


def _mask_cell(cell: str) -> str:
    return "[masked: secret-like cell]" if _SECRET_LINE.search(cell) else cell


def _csv_facts(text: str, delimiter: str) -> Tuple[List[str], List[str]]:
    # 원문(text)을 먼저 파싱하고 셀 단위로 마스킹한다 — 행 전체를 미리 한 줄로
    # 치환하면 'password:' 포함 행이 1셀로 뭉개져 행/열 구조가 왜곡된다.
    rows = [[_mask_cell(cell) for cell in r]
            for r in csv.reader(io.StringIO(text), delimiter=delimiter)
            if any(cell.strip() for cell in r)]
    if not rows:
        return (["빈 CSV"], [])
    header, data = rows[0], rows[1:]
    facts = [f"CSV {len(data)}행 × {len(header)}열 "
             f"(컬럼: {', '.join(h.strip() for h in header)})"]
    notable = []
    for row in data:
        joined = " ".join(row).lower()
        if any(marker in joined for marker in _ABNORMAL_MARKERS):
            notable.append(" / ".join(c.strip() for c in row if c.strip())[:120])
    if notable:
        facts.append(f"이상/주의 행 {len(notable)}건 감지")
    return facts, notable


def _xlsx_facts(path: Path, max_rows: int = 2000) -> Tuple[List[str], List[str], str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)  # type: ignore[union-attr]
    try:
        ws = wb.worksheets[0]
        sheet_names = wb.sheetnames
        rows = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx > max_rows:
                break
            # 다른 입력 경로와 동일하게 비밀번호/토큰 등 secret-like 셀은 마스킹한다.
            values = ["" if cell is None else _mask_secrets(str(cell).strip()) for cell in row]
            if any(values):
                rows.append(values)
        if not rows:
            facts = [f"XLSX 첫 시트 '{ws.title}' 빈 시트"]
            if len(sheet_names) > 1:
                facts.append(f"시트 {len(sheet_names)}개: {', '.join(sheet_names[:8])}")
            return facts, [], ""
        header, data = rows[0], rows[1:]
        facts = [f"XLSX 첫 시트 '{ws.title}' {len(data)}행 × {len(header)}열 "
                 f"(컬럼: {', '.join(h for h in header if h)})"]
        if len(sheet_names) > 1:
            facts.append(f"시트 {len(sheet_names)}개: {', '.join(sheet_names[:8])}")
        if ws.max_row and ws.max_row > max_rows:
            facts.append(f"대용량 XLSX: 앞 {max_rows:,}행만 분석 (전체 {ws.max_row:,}행)")
        notable = []
        for row in data:
            joined = " ".join(row).lower()
            if any(marker in joined for marker in _ABNORMAL_MARKERS):
                notable.append(" / ".join(c for c in row if c)[:120])
        if notable:
            facts.append(f"이상/주의 행 {len(notable)}건 감지")
        preview = "\n".join(" | ".join(c for c in row if c) for row in rows[:8])[:_PREVIEW_CHARS]
        return facts, notable, preview
    finally:
        # 빈 시트 early return·중간 예외 포함 모든 경로에서 핸들 반납 —
        # read_only openpyxl은 파일을 잡고 있어 Windows에서 원본이 잠긴다.
        try:
            wb.close()
        except Exception:
            pass


def _log_facts(text: str) -> Tuple[List[str], List[str]]:
    lines = text.splitlines()
    upper = [line.upper() for line in lines]
    err = sum(1 for line in upper if "ERROR" in line)
    warn = sum(1 for line in upper if "WARN" in line)
    facts = [f"로그 {len(lines)}줄 — ERROR {err}건, WARNING {warn}건"]
    notable = [line.strip()[:120] for line in lines if "ERROR" in line.upper()][:5]
    return facts, notable


def _code_facts(text: str) -> Tuple[List[str], List[str]]:
    entries = _ENTRY_RE.findall(text)
    fact = f"코드 {len(text.splitlines())}줄"
    if entries:
        fact += f" — 진입점: {', '.join(entries[:8])}"
    return [fact], []


def _json_facts(text: str) -> Tuple[List[str], List[str], Optional[List[Dict[str, str]]]]:
    obj = json.loads(text)
    if (isinstance(obj, list) and obj and all(isinstance(x, dict) for x in obj)
            and ("subject" in obj[0] or "제목" in obj[0])):
        mails = [{"from": str(x.get("from", x.get("발신", ""))),
                  "subject": str(x.get("subject", x.get("제목", ""))),
                  "body": str(x.get("body", x.get("내용", "")))} for x in obj]
        return [f"메일 목록 {len(mails)}건 (JSON)"], [], mails
    if isinstance(obj, list):
        return [f"JSON 배열 {len(obj)}개 항목"], [], None
    if isinstance(obj, dict):
        return [f"JSON 객체 (키: {', '.join(list(obj)[:8])})"], [], None
    return ["JSON 스칼라 값"], [], None


def _text_facts(text: str) -> Tuple[List[str], List[str]]:
    lines = [line for line in text.splitlines() if line.strip()]
    return [f"텍스트 {len(lines)}줄"], []


def _preview(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return "\n".join(lines)[:_PREVIEW_CHARS]


def _converted_doc_record(path: Path, suffix: str, conv: Dict[str, Any], size: int) -> Dict[str, Any]:
    text = _mask_secrets(conv.get("markdown", "")[:MAX_BYTES_FULL])
    facts, notable = _text_facts(text)
    facts.insert(0, f"{suffix[1:].upper()} 문서를 markitdown 으로 Markdown 변환 ({conv.get('chars', 0):,}자)")
    return {"name": path.name, "path": str(path), "size_bytes": size,
            "type": f"doc:{suffix[1:]}", "facts": facts,
            "notable": [f"{path.name}: {n}" for n in notable],
            "preview": _preview(text)}


def ingest_inputs(paths: Sequence[Any], max_files: int = MAX_FILES) -> Dict[str, Any]:
    """Read user-supplied files/folders into a secret-free grounding summary.

    Returns {"ok", "files": [{name, path, size_bytes, type, facts, notable,
    preview}], "unsupported": [{name, reason}], "errors": [...],
    "facts": [...], "notable_items": [...], "mails": [...] | None, "summary"}.
    Never raises for a bad path — problems are recorded so the caller (and
    the artifact) can state honestly what was and was not read.
    """
    files: List[Dict[str, Any]] = []
    unsupported: List[Dict[str, str]] = []
    errors: List[str] = []
    mails_all: List[Dict[str, str]] = []
    resolved: List[Path] = []
    truncated = False
    for raw_path in paths:
        path = Path(str(raw_path)).expanduser()
        if not path.exists():
            errors.append(f"입력 경로 없음: {path}")
            continue
        if path.is_dir():
            # max_files개를 모으면 즉시 중단 — 전체 트리를 is_file()로 훑지 않아
            # OneDrive 클라우드 placeholder 파일의 대량 하이드레이션을 막는다.
            collected: List[Path] = []
            for p in path.rglob("*"):
                if len(resolved) + len(collected) >= max_files:
                    truncated = True
                    break
                if p.is_file():
                    collected.append(p)
            resolved.extend(sorted(collected))
        elif len(resolved) >= max_files:
            truncated = True
        else:
            resolved.append(path)
    if truncated:
        errors.append(f"입력 파일이 많아 앞 {max_files}개만 분석 (초과분은 건너뜀)")
    for path in resolved:
        suffix = path.suffix.lower()
        # PDF/DOCX/PPTX/HTML 등은 markitdown(오프라인)으로 Markdown 변환 후 텍스트 분석.
        if suffix not in SUPPORTED_SUFFIXES and doc_convert.can_convert(suffix):
            conv = doc_convert.convert_file(path)
            if not conv.get("ok"):
                unsupported.append({"name": path.name,
                                    "reason": conv.get("error", "문서 변환 불가")
                                    + (f" — {conv['hint']}" if conv.get("hint") else "")})
                continue
            try:
                size = path.stat().st_size
                files.append(_converted_doc_record(path, suffix, conv, size))
            except Exception as exc:  # noqa: BLE001
                errors.append(f"{path.name}: 문서 분석 실패 {exc!r}"[:200])
            continue
        if suffix not in SUPPORTED_SUFFIXES:
            hint = ""
            if doc_convert.can_convert(suffix):  # 도달 안 하지만 방어적
                hint = f" — {doc_convert._hint_for(suffix)} 반입 시 지원"
            unsupported.append({"name": path.name,
                                "reason": f"확장자 {suffix or '(없음)'} 미지원 — 텍스트 계열(MD/TXT/CSV/LOG/PY/BAS/BAT/JSON)만 직접 읽음"
                                          "; 문서(PDF/DOCX/PPTX/HTML)는 markitdown 반입 필요" + hint})
            continue
        if suffix == ".xlsx":
            if openpyxl is None:
                if doc_convert.can_convert(suffix):
                    conv = doc_convert.convert_file(path)
                    if conv.get("ok"):
                        try:
                            files.append(_converted_doc_record(path, suffix, conv, path.stat().st_size))
                        except Exception as exc:  # noqa: BLE001
                            errors.append(f"{path.name}: 문서 분석 실패 {exc!r}"[:200])
                    else:
                        unsupported.append({"name": path.name,
                                            "reason": "openpyxl 미설치; "
                                            + conv.get("error", "markitdown 변환 불가")
                                            + (f" — {conv['hint']}" if conv.get("hint") else "")})
                else:
                    unsupported.append({"name": path.name,
                                        "reason": "openpyxl 미설치 (dependencies.json 'office-doc-wheels')"})
                continue
            try:
                size = path.stat().st_size
                facts, notable, preview = _xlsx_facts(path)
            except Exception as exc:
                errors.append(f"{path.name}: XLSX 읽기 실패 {exc!r}"[:200])
                continue
            files.append({"name": path.name, "path": str(path), "size_bytes": size,
                          "type": "xlsx", "facts": facts,
                          "notable": [f"{path.name}: {n}" for n in notable],
                          "preview": preview})
            continue
        try:
            size = path.stat().st_size
            with path.open("rb") as fh:
                raw = fh.read(MAX_BYTES_FULL)
            text = decode_file_bytes(raw, truncated=size > MAX_BYTES_FULL)
        except Exception as exc:
            errors.append(f"{path.name}: 읽기 실패 {exc!r}"[:200])
            continue
        if "\x00" in text[:2000]:
            unsupported.append({"name": path.name, "reason": "binary content (텍스트 아님)"})
            continue
        # CSV/TSV는 파싱 전에 마스킹하면 행/열 구조가 깨지므로 원문을 넘기고
        # _csv_facts 내부에서 셀 단위 마스킹한다. preview용 text는 아래에서 마스킹.
        raw_text = text
        text = _mask_secrets(text)
        mails: Optional[List[Dict[str, str]]] = None
        try:
            if suffix in (".csv", ".tsv"):
                facts, notable = _csv_facts(raw_text, "\t" if suffix == ".tsv" else ",")
                kind = "csv"
            elif suffix == ".log":
                facts, notable = _log_facts(text)
                kind = "log"
            elif suffix in (".py", ".bas", ".bat"):
                facts, notable = _code_facts(text)
                kind = "code"
            elif suffix == ".json":
                facts, notable, mails = _json_facts(text)
                kind = "mail_list" if mails else "json"
            else:
                facts, notable = _text_facts(text)
                kind = "text"
        except Exception as exc:
            facts, notable, kind = [f"내용 분석 실패: {exc!r}"[:120]], [], "text"
        if size > MAX_BYTES_FULL:
            facts.append(f"파일이 큼: {size:,} bytes 중 앞 {MAX_BYTES_FULL:,} bytes만 분석")
        files.append({"name": path.name, "path": str(path), "size_bytes": size,
                      "type": kind, "facts": facts,
                      "notable": [f"{path.name}: {n}" for n in notable],
                      "preview": _preview(text)})
        if mails:
            mails_all.extend(mails)
    return {
        "ok": bool(files),
        "files": files,
        "unsupported": unsupported,
        "errors": errors,
        "facts": [f"{f['name']}: {fact}" for f in files for fact in f["facts"]],
        "notable_items": [n for f in files for n in f["notable"]],
        "mails": mails_all or None,
        "summary": (f"입력 {len(files)}건 분석"
                    + (f", 미지원 {len(unsupported)}건" if unsupported else "")
                    + (f", 오류 {len(errors)}건" if errors else "")),
    }
